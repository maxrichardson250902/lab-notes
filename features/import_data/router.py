"""DNA Manager feature — primers, plasmids, gBlocks, kit parts, storage boxes, .gb files, and auto-linking."""

from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
from core.database import register_table, get_db
import csv
import io
import os
import uuid
import re
import json

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from Bio import SeqIO
    HAS_BIOPYTHON = True
except ImportError:
    HAS_BIOPYTHON = False

# ── Tables ───────────────────────────────────────────────────────────────────

register_table("imports", """CREATE TABLE IF NOT EXISTS imports (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    filename     TEXT NOT NULL,
    record_type  TEXT NOT NULL,
    record_count INTEGER NOT NULL DEFAULT 0,
    created      TEXT NOT NULL)""")

register_table("primers", """CREATE TABLE IF NOT EXISTS primers (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id  INTEGER,
    name       TEXT NOT NULL,
    sequence   TEXT,
    use        TEXT,
    box_number TEXT,
    gb_file    TEXT,
    project    TEXT DEFAULT '',
    created    TEXT NOT NULL)""")

register_table("plasmids", """CREATE TABLE IF NOT EXISTS plasmids (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id         INTEGER,
    name              TEXT NOT NULL,
    use               TEXT,
    box_location      TEXT,
    glycerol_location TEXT,
    gb_file           TEXT,
    project           TEXT DEFAULT '',
    created           TEXT NOT NULL)""")

register_table("dna_settings", """CREATE TABLE IF NOT EXISTS dna_settings (
    id              INTEGER PRIMARY KEY CHECK (id = 1),
    primer_prefix   TEXT NOT NULL DEFAULT '',
    plasmid_prefix  TEXT NOT NULL DEFAULT '',
    created         TEXT NOT NULL)""")

register_table("gblocks", """CREATE TABLE IF NOT EXISTS gblocks (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id   INTEGER,
    name        TEXT NOT NULL,
    sequence    TEXT,
    length      INTEGER,
    project     TEXT DEFAULT '',
    use         TEXT,
    supplier    TEXT DEFAULT 'IDT',
    order_id    TEXT,
    box_number  TEXT,
    gb_file     TEXT,
    notes       TEXT,
    created     TEXT NOT NULL)""")

register_table("kit_parts", """CREATE TABLE IF NOT EXISTS kit_parts (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id         INTEGER,
    name              TEXT NOT NULL,
    kit_name          TEXT,
    part_type         TEXT,
    description       TEXT,
    project           TEXT DEFAULT '',
    resistance        TEXT,
    box_location      TEXT,
    glycerol_location TEXT,
    gb_file           TEXT,
    source_url        TEXT,
    notes             TEXT,
    created           TEXT NOT NULL)""")

register_table("parts", """CREATE TABLE IF NOT EXISTS parts (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    import_id         INTEGER,
    name              TEXT NOT NULL,
    description       TEXT,
    sequence          TEXT,
    length            INTEGER,
    project           TEXT DEFAULT '',
    subcategory       TEXT DEFAULT '',
    source_feature    TEXT,
    source_id         TEXT,
    part_type         TEXT,
    box_location      TEXT,
    glycerol_location TEXT,
    gb_file           TEXT,
    notes             TEXT,
    created           TEXT NOT NULL)""")

register_table("storage_boxes", """CREATE TABLE IF NOT EXISTS storage_boxes (
    id       INTEGER PRIMARY KEY AUTOINCREMENT,
    name     TEXT NOT NULL,
    rows     INTEGER NOT NULL DEFAULT 9,
    cols     INTEGER NOT NULL DEFAULT 9,
    box_type TEXT DEFAULT 'mixed',
    location TEXT,
    layout   TEXT DEFAULT '{}',
    created  TEXT NOT NULL,
    updated  TEXT NOT NULL)""")

# ── Storage directories ──────────────────────────────────────────────────────

UPLOAD_DIR = "/data/imports"
GB_DIR = "/data/gb_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(GB_DIR, exist_ok=True)

# ── Antibiotic resistance detection ─────────────────────────────────────────

RESISTANCE_GENES = {
    r'\bbla\b': 'Ampicillin',
    r'\bampR\b': 'Ampicillin',
    r'\bamp\b': 'Ampicillin',
    r'beta-lactamase': 'Ampicillin',
    r'\bkan\b': 'Kanamycin',
    r'\bkanR\b': 'Kanamycin',
    r'\baph\b': 'Kanamycin',
    r'\bneo\b': 'Kanamycin/Neomycin',
    r'\bnptII\b': 'Kanamycin',
    r'\bcat\b': 'Chloramphenicol',
    r'\bcmR\b': 'Chloramphenicol',
    r'chloramphenicol acetyltransferase': 'Chloramphenicol',
    r'\btet\b': 'Tetracycline',
    r'\btetR\b': 'Tetracycline',
    r'\btetA\b': 'Tetracycline',
    r'\bspec\b': 'Spectinomycin',
    r'\baadA\b': 'Spectinomycin',
    r'\berm\b': 'Erythromycin',
    r'\bhyg\b': 'Hygromycin',
    r'\bhygR\b': 'Hygromycin',
    r'\bhph\b': 'Hygromycin',
    r'\bzeo\b': 'Zeocin',
    r'\bble\b': 'Zeocin',
    r'\bbsr\b': 'Blasticidin',
    r'\bpac\b': 'Puromycin',
    r'\bpuro\b': 'Puromycin',
    r'\bnat\b': 'Nourseothricin',
    r'\bgent\b': 'Gentamicin',
    r'\baac': 'Gentamicin',
    r'\bstrep\b': 'Streptomycin',
    r'\bsulI\b': 'Sulfonamide',
    r'\btmp\b': 'Trimethoprim',
    r'\bdhfr\b': 'Trimethoprim',
}

RESISTANCE_KEYWORDS = [
    'ampicillin', 'kanamycin', 'chloramphenicol', 'tetracycline',
    'spectinomycin', 'erythromycin', 'hygromycin', 'zeocin',
    'blasticidin', 'puromycin', 'nourseothricin', 'gentamicin',
    'streptomycin', 'sulfonamide', 'trimethoprim', 'carbenicillin',
    'neomycin', 'G418',
]

RESISTANCE_SHORTHANDS = {
    'AmpR': 'Ampicillin', 'Amp(R)': 'Ampicillin',
    'KanR': 'Kanamycin', 'Kan(R)': 'Kanamycin',
    'CmR': 'Chloramphenicol', 'Cm(R)': 'Chloramphenicol',
    'TetR': 'Tetracycline', 'Tet(R)': 'Tetracycline',
    'SpecR': 'Spectinomycin', 'Spec(R)': 'Spectinomycin',
    'ErmR': 'Erythromycin',
    'HygR': 'Hygromycin', 'Hyg(R)': 'Hygromycin',
    'ZeoR': 'Zeocin', 'Zeo(R)': 'Zeocin',
    'BsrR': 'Blasticidin', 'BlastR': 'Blasticidin',
    'PuroR': 'Puromycin', 'Puro(R)': 'Puromycin',
    'NatR': 'Nourseothricin',
    'GenR': 'Gentamicin', 'Gen(R)': 'Gentamicin',
    'StrepR': 'Streptomycin',
    'NeoR': 'Kanamycin/Neomycin', 'Neo(R)': 'Kanamycin/Neomycin',
    'SmR': 'Streptomycin',
}


def _parse_resistance_from_gb(contents: bytes) -> str:
    if not HAS_BIOPYTHON:
        return ""
    found = set()
    try:
        records = list(SeqIO.parse(io.StringIO(contents.decode("utf-8", errors="replace")), "genbank"))
    except Exception:
        return ""
    for record in records:
        for feature in record.features:
            all_text = ""
            for key, vals in feature.qualifiers.items():
                for v in vals:
                    all_text += " " + v
            lower_text = all_text.lower()
            for kw in RESISTANCE_KEYWORDS:
                if kw.lower() in lower_text and any(
                    t in lower_text for t in ("resistance", "resistant", "selectable marker")
                ):
                    found.add(kw.capitalize())
            for pattern, antibiotic in RESISTANCE_GENES.items():
                if re.search(pattern, all_text, re.IGNORECASE):
                    found.add(antibiotic)
            for shorthand, antibiotic in RESISTANCE_SHORTHANDS.items():
                if re.search(r'(?:^|\s|")' + re.escape(shorthand) + r'(?:\s|"|$)', all_text, re.IGNORECASE):
                    found.add(antibiotic)
            for kw in RESISTANCE_KEYWORDS:
                if kw.lower() in lower_text:
                    found.add(kw.capitalize())
    normalized = set()
    for a in found:
        a_lower = a.lower()
        if a_lower == 'g418':
            normalized.add('Kanamycin/G418')
        elif a_lower == 'carbenicillin':
            normalized.add('Ampicillin/Carbenicillin')
        elif a_lower == 'neomycin':
            normalized.add('Kanamycin/Neomycin')
        else:
            normalized.add(a)
    return ", ".join(sorted(normalized))


# ── Router ───────────────────────────────────────────────────────────────────

router = APIRouter(prefix="/api", tags=["dna_manager"])

# ── Ensure migration columns exist (lazy) ───────────────────────────────────

_migration_checked = False

def _ensure_migrations():
    global _migration_checked
    if _migration_checked:
        return
    _migration_checked = True
    with get_db() as conn:
        # plasmids: antibiotic_resistance, project, subcategory
        pcols = [row[1] for row in conn.execute("PRAGMA table_info(plasmids)").fetchall()]
        if "antibiotic_resistance" not in pcols:
            conn.execute("ALTER TABLE plasmids ADD COLUMN antibiotic_resistance TEXT DEFAULT ''")
        if "project" not in pcols:
            conn.execute("ALTER TABLE plasmids ADD COLUMN project TEXT DEFAULT ''")
        if "subcategory" not in pcols:
            conn.execute("ALTER TABLE plasmids ADD COLUMN subcategory TEXT DEFAULT ''")
        # primers: project, subcategory
        prcols = [row[1] for row in conn.execute("PRAGMA table_info(primers)").fetchall()]
        if "project" not in prcols:
            conn.execute("ALTER TABLE primers ADD COLUMN project TEXT DEFAULT ''")
        if "subcategory" not in prcols:
            conn.execute("ALTER TABLE primers ADD COLUMN subcategory TEXT DEFAULT ''")
        # gblocks: subcategory
        try:
            gcols = [row[1] for row in conn.execute("PRAGMA table_info(gblocks)").fetchall()]
            if "subcategory" not in gcols:
                conn.execute("ALTER TABLE gblocks ADD COLUMN subcategory TEXT DEFAULT ''")
        except Exception:
            pass
        # kit_parts: subcategory
        try:
            kcols = [row[1] for row in conn.execute("PRAGMA table_info(kit_parts)").fetchall()]
            if "subcategory" not in kcols:
                conn.execute("ALTER TABLE kit_parts ADD COLUMN subcategory TEXT DEFAULT ''")
        except Exception:
            pass
        # Stock status columns on all DNA tables
        for _tbl in ("primers", "plasmids", "gblocks", "kit_parts", "parts"):
            try:
                _cols = [row[1] for row in conn.execute(f"PRAGMA table_info({_tbl})").fetchall()]
                for _sc in ("stock_dna", "stock_glycerol", "stock_verified"):
                    if _sc not in _cols:
                        conn.execute(f"ALTER TABLE {_tbl} ADD COLUMN {_sc} TEXT DEFAULT ''")
            except Exception:
                pass
        conn.commit()


# ══════════════════════════════════════════════════════════════════════════════
#  PROJECTS HELPER
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/dna/projects")
def list_projects():
    _ensure_migrations()
    projects = set()
    subcategories = {}  # project -> set of subcategories
    with get_db() as conn:
        for tbl in ("primers", "plasmids", "gblocks", "kit_parts", "parts"):
            try:
                rows = conn.execute(f"SELECT DISTINCT project, subcategory FROM {tbl} WHERE project IS NOT NULL AND project != ''").fetchall()
                for r in rows:
                    proj = r["project"]
                    sub = r["subcategory"] or ""
                    projects.add(proj)
                    if sub:
                        if proj not in subcategories:
                            subcategories[proj] = set()
                        subcategories[proj].add(sub)
            except Exception:
                pass
    return {
        "projects": sorted(projects),
        "subcategories": {k: sorted(v) for k, v in subcategories.items()}
    }


# ══════════════════════════════════════════════════════════════════════════════
#  MOVE DNA BETWEEN TYPES
# ══════════════════════════════════════════════════════════════════════════════

class MoveDnaRequest(BaseModel):
    from_type: str    # primer, plasmid, gblock, kit_part
    from_id: int
    to_type: str      # primer, plasmid, gblock, kit_part

# Map type -> (table, file_prefix)
_TYPE_MAP = {
    "primer":   ("primers",   "primer"),
    "plasmid":  ("plasmids",  "plasmid"),
    "gblock":   ("gblocks",   "gblock"),
    "kit_part": ("kit_parts", "kitpart"),
    "part":     ("parts",     "part"),
}

@router.post("/dna/move")
def move_dna(body: MoveDnaRequest):
    _ensure_migrations()
    if body.from_type not in _TYPE_MAP or body.to_type not in _TYPE_MAP:
        raise HTTPException(400, f"Types must be one of {list(_TYPE_MAP.keys())}")
    if body.from_type == body.to_type:
        raise HTTPException(400, "Source and destination types are the same")

    src_table, src_prefix = _TYPE_MAP[body.from_type]
    dst_table, dst_prefix = _TYPE_MAP[body.to_type]

    now = datetime.utcnow().isoformat()
    with get_db() as conn:
        row = conn.execute(f"SELECT * FROM {src_table} WHERE id=?", (body.from_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Source item not found")
        src = dict(row)

        # Common fields
        name = src.get("name", "")
        project = src.get("project", "")
        subcategory = src.get("subcategory", "")
        use = src.get("use", "") or src.get("description", "")
        sequence = src.get("sequence", "")
        gb_file = src.get("gb_file", "")

        # Insert into destination
        if body.to_type == "primer":
            cur = conn.execute(
                "INSERT INTO primers (name, sequence, use, box_number, project, subcategory, created) VALUES (?,?,?,?,?,?,?)",
                (name, sequence, use, src.get("box_number", "") or src.get("box_location", ""), project, subcategory, now))
        elif body.to_type == "plasmid":
            cur = conn.execute(
                "INSERT INTO plasmids (name, use, box_location, glycerol_location, project, subcategory, created) VALUES (?,?,?,?,?,?,?)",
                (name, use, src.get("box_location", "") or src.get("box_number", ""),
                 src.get("glycerol_location", ""), project, subcategory, now))
        elif body.to_type == "gblock":
            length = len(re.sub(r'[^ACGTacgt]', '', sequence)) if sequence else 0
            cur = conn.execute(
                "INSERT INTO gblocks (name, sequence, length, project, subcategory, use, supplier, box_number, notes, created) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (name, sequence, length, project, subcategory, use, src.get("supplier", ""),
                 src.get("box_number", "") or src.get("box_location", ""), src.get("notes", ""), now))
        elif body.to_type == "kit_part":
            cur = conn.execute(
                "INSERT INTO kit_parts (name, description, project, subcategory, resistance, box_location, "
                "glycerol_location, notes, created) VALUES (?,?,?,?,?,?,?,?,?)",
                (name, use, project, subcategory, src.get("resistance", "") or src.get("antibiotic_resistance", ""),
                 src.get("box_location", "") or src.get("box_number", ""),
                 src.get("glycerol_location", ""), src.get("notes", ""), now))
        elif body.to_type == "part":
            length = len(re.sub(r'[^ACGTacgt]', '', sequence)) if sequence else 0
            cur = conn.execute(
                "INSERT INTO parts (name, description, sequence, length, project, subcategory, source_feature, "
                "part_type, box_location, glycerol_location, notes, created) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (name, use, sequence, length, project, subcategory,
                 src.get("source_feature", ""), src.get("part_type", ""),
                 src.get("box_location", "") or src.get("box_number", ""),
                 src.get("glycerol_location", ""), src.get("notes", ""), now))

        new_id = cur.lastrowid

        # Move .gb file if it exists
        if gb_file:
            old_path = os.path.join(GB_DIR, f"{src_prefix}_{body.from_id}.gb")
            new_path = os.path.join(GB_DIR, f"{dst_prefix}_{new_id}.gb")
            if os.path.exists(old_path):
                import shutil
                shutil.copy2(old_path, new_path)
                os.remove(old_path)
            conn.execute(f"UPDATE {dst_table} SET gb_file=? WHERE id=?", (gb_file, new_id))

        # Delete source
        conn.execute(f"DELETE FROM {src_table} WHERE id=?", (body.from_id,))
        conn.commit()

    return {"ok": True, "new_type": body.to_type, "new_id": new_id}


# ══════════════════════════════════════════════════════════════════════════════
#  PRIMERS CRUD
# ══════════════════════════════════════════════════════════════════════════════

class CreatePrimer(BaseModel):
    name: str
    sequence: Optional[str] = ""
    use: Optional[str] = ""
    box_number: Optional[str] = ""
    project: Optional[str] = ""
    subcategory: Optional[str] = ""

class UpdatePrimer(BaseModel):
    name: Optional[str] = None
    sequence: Optional[str] = None
    use: Optional[str] = None
    box_number: Optional[str] = None
    project: Optional[str] = None
    subcategory: Optional[str] = None
    stock_dna: Optional[str] = None
    stock_glycerol: Optional[str] = None
    stock_verified: Optional[str] = None

@router.get("/primers")
def list_primers():
    _ensure_migrations()
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM primers ORDER BY name ASC").fetchall()
    return {"items": [dict(r) for r in rows]}

@router.post("/primers")
def create_primer(body: CreatePrimer):
    _ensure_migrations()
    now = datetime.utcnow().isoformat()
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO primers (name, sequence, use, box_number, project, subcategory, created) VALUES (?,?,?,?,?,?,?)",
            (body.name, body.sequence, body.use, body.box_number, body.project, body.subcategory, now))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM primers WHERE id=?", (cur.lastrowid,)).fetchone())
    return row

@router.put("/primers/{item_id}")
def update_primer(item_id: int, body: UpdatePrimer):
    _ensure_migrations()
    with get_db() as conn:
        existing = conn.execute("SELECT * FROM primers WHERE id=?", (item_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Primer not found")
        fields = {k: v for k, v in body.dict().items() if v is not None}
        if not fields:
            return dict(existing)
        sets = ", ".join(f"{k}=?" for k in fields)
        conn.execute(f"UPDATE primers SET {sets} WHERE id=?", (*fields.values(), item_id))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM primers WHERE id=?", (item_id,)).fetchone())
    return row

@router.delete("/primers/{item_id}")
def delete_primer(item_id: int):
    path = os.path.join(GB_DIR, f"primer_{item_id}.gb")
    if os.path.exists(path):
        os.remove(path)
    with get_db() as conn:
        conn.execute("DELETE FROM primers WHERE id=?", (item_id,))
        conn.commit()
    return {"ok": True}

# ── Primer .gb file ──────────────────────────────────────────────────────────

@router.post("/primers/{item_id}/gb")
async def upload_primer_gb(item_id: int, file: UploadFile = File(...)):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM primers WHERE id=?", (item_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Primer not found")
    contents = await file.read()
    stored = os.path.join(GB_DIR, f"primer_{item_id}.gb")
    with open(stored, "wb") as f:
        f.write(contents)
    with get_db() as conn:
        conn.execute("UPDATE primers SET gb_file=? WHERE id=?", (file.filename, item_id))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM primers WHERE id=?", (item_id,)).fetchone())
    return row

@router.get("/primers/{item_id}/gb")
def download_primer_gb(item_id: int):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM primers WHERE id=?", (item_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Primer not found")
    stored = os.path.join(GB_DIR, f"primer_{item_id}.gb")
    if not os.path.exists(stored):
        raise HTTPException(404, "No .gb file attached")
    return FileResponse(stored, filename=row["gb_file"] or f"primer_{item_id}.gb",
                        media_type="application/octet-stream")

@router.delete("/primers/{item_id}/gb")
def delete_primer_gb(item_id: int):
    stored = os.path.join(GB_DIR, f"primer_{item_id}.gb")
    if os.path.exists(stored):
        os.remove(stored)
    with get_db() as conn:
        conn.execute("UPDATE primers SET gb_file=NULL WHERE id=?", (item_id,))
        conn.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
#  PLASMIDS CRUD
# ══════════════════════════════════════════════════════════════════════════════

class CreatePlasmid(BaseModel):
    name: str
    use: Optional[str] = ""
    box_location: Optional[str] = ""
    glycerol_location: Optional[str] = ""
    project: Optional[str] = ""
    subcategory: Optional[str] = ""

class UpdatePlasmid(BaseModel):
    name: Optional[str] = None
    use: Optional[str] = None
    box_location: Optional[str] = None
    glycerol_location: Optional[str] = None
    antibiotic_resistance: Optional[str] = None
    project: Optional[str] = None
    subcategory: Optional[str] = None
    stock_dna: Optional[str] = None
    stock_glycerol: Optional[str] = None
    stock_verified: Optional[str] = None

@router.get("/plasmids")
def list_plasmids():
    _ensure_migrations()
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM plasmids ORDER BY name ASC").fetchall()
    return {"items": [dict(r) for r in rows]}

@router.post("/plasmids")
def create_plasmid(body: CreatePlasmid):
    _ensure_migrations()
    now = datetime.utcnow().isoformat()
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO plasmids (name, use, box_location, glycerol_location, project, subcategory, created) VALUES (?,?,?,?,?,?,?)",
            (body.name, body.use, body.box_location, body.glycerol_location, body.project, body.subcategory, now))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM plasmids WHERE id=?", (cur.lastrowid,)).fetchone())
    return row

@router.put("/plasmids/{item_id}")
def update_plasmid(item_id: int, body: UpdatePlasmid):
    _ensure_migrations()
    with get_db() as conn:
        existing = conn.execute("SELECT * FROM plasmids WHERE id=?", (item_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Plasmid not found")
        fields = {k: v for k, v in body.dict().items() if v is not None}
        if not fields:
            return dict(existing)
        sets = ", ".join(f"{k}=?" for k in fields)
        conn.execute(f"UPDATE plasmids SET {sets} WHERE id=?", (*fields.values(), item_id))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM plasmids WHERE id=?", (item_id,)).fetchone())
    return row

@router.delete("/plasmids/{item_id}")
def delete_plasmid(item_id: int):
    path = os.path.join(GB_DIR, f"plasmid_{item_id}.gb")
    if os.path.exists(path):
        os.remove(path)
    with get_db() as conn:
        conn.execute("DELETE FROM plasmids WHERE id=?", (item_id,))
        conn.commit()
    return {"ok": True}

# ── Plasmid .gb file ─────────────────────────────────────────────────────────

@router.post("/plasmids/{item_id}/gb")
async def upload_plasmid_gb(item_id: int, file: UploadFile = File(...)):
    _ensure_migrations()
    with get_db() as conn:
        row = conn.execute("SELECT * FROM plasmids WHERE id=?", (item_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Plasmid not found")
    contents = await file.read()
    stored = os.path.join(GB_DIR, f"plasmid_{item_id}.gb")
    with open(stored, "wb") as f:
        f.write(contents)
    resistance = _parse_resistance_from_gb(contents)
    with get_db() as conn:
        conn.execute(
            "UPDATE plasmids SET gb_file=?, antibiotic_resistance=? WHERE id=?",
            (file.filename, resistance, item_id))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM plasmids WHERE id=?", (item_id,)).fetchone())
    return row

@router.get("/plasmids/{item_id}/gb")
def download_plasmid_gb(item_id: int):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM plasmids WHERE id=?", (item_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Plasmid not found")
    stored = os.path.join(GB_DIR, f"plasmid_{item_id}.gb")
    if not os.path.exists(stored):
        raise HTTPException(404, "No .gb file attached")
    return FileResponse(stored, filename=row["gb_file"] or f"plasmid_{item_id}.gb",
                        media_type="application/octet-stream")

@router.delete("/plasmids/{item_id}/gb")
def delete_plasmid_gb(item_id: int):
    _ensure_migrations()
    stored = os.path.join(GB_DIR, f"plasmid_{item_id}.gb")
    if os.path.exists(stored):
        os.remove(stored)
    with get_db() as conn:
        conn.execute("UPDATE plasmids SET gb_file=NULL, antibiotic_resistance='' WHERE id=?", (item_id,))
        conn.commit()
    return {"ok": True}

# ── Re-scan .gb files for resistance ─────────────────────────────────────────

@router.post("/plasmids/rescan-resistance")
def rescan_all_resistance():
    _ensure_migrations()
    updated = 0
    with get_db() as conn:
        rows = conn.execute("SELECT id, gb_file FROM plasmids WHERE gb_file IS NOT NULL AND gb_file != ''").fetchall()
        for row in rows:
            stored = os.path.join(GB_DIR, f"plasmid_{row['id']}.gb")
            if not os.path.exists(stored):
                continue
            with open(stored, "rb") as f:
                contents = f.read()
            resistance = _parse_resistance_from_gb(contents)
            conn.execute("UPDATE plasmids SET antibiotic_resistance=? WHERE id=?", (resistance, row["id"]))
            updated += 1
        conn.commit()
    return {"updated": updated}


# ══════════════════════════════════════════════════════════════════════════════
#  GBLOCKS CRUD
# ══════════════════════════════════════════════════════════════════════════════

class CreateGblock(BaseModel):
    name: str
    sequence: Optional[str] = ""
    project: Optional[str] = ""
    subcategory: Optional[str] = ""
    use: Optional[str] = ""
    supplier: Optional[str] = "IDT"
    order_id: Optional[str] = ""
    box_number: Optional[str] = ""
    notes: Optional[str] = ""

class UpdateGblock(BaseModel):
    name: Optional[str] = None
    sequence: Optional[str] = None
    project: Optional[str] = None
    subcategory: Optional[str] = None
    use: Optional[str] = None
    supplier: Optional[str] = None
    order_id: Optional[str] = None
    box_number: Optional[str] = None
    notes: Optional[str] = None
    stock_dna: Optional[str] = None
    stock_glycerol: Optional[str] = None
    stock_verified: Optional[str] = None

def _calc_gblock_length(seq: str) -> int:
    if not seq:
        return 0
    return len(re.sub(r'[^ACGTacgt]', '', seq))


def _extract_seq_from_gb(contents: bytes) -> str:
    """Extract the DNA sequence from a GenBank file."""
    if not HAS_BIOPYTHON:
        return ""
    try:
        from Bio import SeqIO
        records = list(SeqIO.parse(io.StringIO(contents.decode("utf-8", errors="replace")), "genbank"))
        if records:
            return str(records[0].seq)
    except Exception:
        pass
    return ""


def _generate_gb_for_gblock(item_id: int, name: str, sequence: str):
    """Generate a minimal GenBank file from a name and sequence, save to disk,
    and return the filename. Returns empty string if BioPython not available or no sequence."""
    if not sequence or not HAS_BIOPYTHON:
        return ""
    clean_seq = re.sub(r'[^ACGTacgtNn]', '', sequence)
    if not clean_seq:
        return ""
    try:
        from Bio.Seq import Seq
        from Bio.SeqRecord import SeqRecord
        from Bio.SeqFeature import SeqFeature, FeatureLocation
        from Bio import SeqIO as _SeqIO

        rec = SeqRecord(
            Seq(clean_seq.upper()),
            id=name[:10].replace(" ", "_"),
            name=name[:10].replace(" ", "_"),
            description=name,
            annotations={"molecule_type": "DNA", "topology": "linear"},
        )
        # Add a source feature
        rec.features.append(SeqFeature(
            FeatureLocation(0, len(clean_seq)),
            type="source",
            qualifiers={"mol_type": ["other DNA"], "organism": ["synthetic construct"],
                        "label": [name]},
        ))
        # Add a misc_feature spanning the whole gBlock
        rec.features.append(SeqFeature(
            FeatureLocation(0, len(clean_seq)),
            type="misc_feature",
            qualifiers={"label": [name], "note": ["gBlock synthetic fragment"]},
        ))

        gb_path = os.path.join(GB_DIR, f"gblock_{item_id}.gb")
        with open(gb_path, "w") as f:
            _SeqIO.write(rec, f, "genbank")

        filename = f"{name}.gb".replace(" ", "_")
        return filename
    except Exception:
        return ""


@router.get("/gblocks")
def list_gblocks():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM gblocks ORDER BY name ASC").fetchall()
    return {"items": [dict(r) for r in rows]}

@router.post("/gblocks")
def create_gblock(body: CreateGblock):
    now = datetime.utcnow().isoformat()
    length = _calc_gblock_length(body.sequence or "")
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO gblocks (name, sequence, length, project, subcategory, use, supplier, order_id, box_number, notes, created) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (body.name, body.sequence, length, body.project, body.subcategory, body.use,
             body.supplier, body.order_id, body.box_number, body.notes, now))
        conn.commit()
        item_id = cur.lastrowid

        # Auto-generate .gb from sequence if sequence provided and no .gb will be uploaded separately
        gb_file = ""
        if body.sequence and body.sequence.strip():
            gb_file = _generate_gb_for_gblock(item_id, body.name, body.sequence)
            if gb_file:
                conn.execute("UPDATE gblocks SET gb_file=? WHERE id=?", (gb_file, item_id))
                conn.commit()

        row = dict(conn.execute("SELECT * FROM gblocks WHERE id=?", (item_id,)).fetchone())
    return row

@router.put("/gblocks/{item_id}")
def update_gblock(item_id: int, body: UpdateGblock):
    with get_db() as conn:
        existing = conn.execute("SELECT * FROM gblocks WHERE id=?", (item_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "gBlock not found")
        fields = {k: v for k, v in body.dict().items() if v is not None}
        if not fields:
            return dict(existing)
        # recalculate length if sequence changed
        if "sequence" in fields:
            fields["length"] = _calc_gblock_length(fields["sequence"])
        sets = ", ".join(f"{k}=?" for k in fields)
        conn.execute(f"UPDATE gblocks SET {sets} WHERE id=?", (*fields.values(), item_id))
        conn.commit()

        # Regenerate .gb if sequence was updated
        if "sequence" in fields and fields["sequence"].strip():
            name = fields.get("name") or dict(existing)["name"]
            gb_file = _generate_gb_for_gblock(item_id, name, fields["sequence"])
            if gb_file:
                conn.execute("UPDATE gblocks SET gb_file=? WHERE id=?", (gb_file, item_id))
                conn.commit()

        row = dict(conn.execute("SELECT * FROM gblocks WHERE id=?", (item_id,)).fetchone())
    return row

@router.delete("/gblocks/{item_id}")
def delete_gblock(item_id: int):
    path = os.path.join(GB_DIR, f"gblock_{item_id}.gb")
    if os.path.exists(path):
        os.remove(path)
    with get_db() as conn:
        conn.execute("DELETE FROM gblocks WHERE id=?", (item_id,))
        conn.commit()
    return {"ok": True}

# ── gBlock .gb file ──────────────────────────────────────────────────────────

@router.post("/gblocks/{item_id}/gb")
async def upload_gblock_gb(item_id: int, file: UploadFile = File(...)):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM gblocks WHERE id=?", (item_id,)).fetchone()
        if not row:
            raise HTTPException(404, "gBlock not found")
    contents = await file.read()
    stored = os.path.join(GB_DIR, f"gblock_{item_id}.gb")
    with open(stored, "wb") as f:
        f.write(contents)

    # Extract sequence from the .gb file
    extracted_seq = _extract_seq_from_gb(contents)
    length = _calc_gblock_length(extracted_seq) if extracted_seq else (dict(row).get("length") or 0)

    with get_db() as conn:
        if extracted_seq:
            conn.execute(
                "UPDATE gblocks SET gb_file=?, sequence=?, length=? WHERE id=?",
                (file.filename, extracted_seq, length, item_id))
        else:
            conn.execute("UPDATE gblocks SET gb_file=? WHERE id=?", (file.filename, item_id))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM gblocks WHERE id=?", (item_id,)).fetchone())
    return row

@router.get("/gblocks/{item_id}/gb")
def download_gblock_gb(item_id: int):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM gblocks WHERE id=?", (item_id,)).fetchone()
        if not row:
            raise HTTPException(404, "gBlock not found")
    stored = os.path.join(GB_DIR, f"gblock_{item_id}.gb")
    if not os.path.exists(stored):
        raise HTTPException(404, "No .gb file attached")
    return FileResponse(stored, filename=row["gb_file"] or f"gblock_{item_id}.gb",
                        media_type="application/octet-stream")

@router.delete("/gblocks/{item_id}/gb")
def delete_gblock_gb(item_id: int):
    stored = os.path.join(GB_DIR, f"gblock_{item_id}.gb")
    if os.path.exists(stored):
        os.remove(stored)
    with get_db() as conn:
        conn.execute("UPDATE gblocks SET gb_file=NULL WHERE id=?", (item_id,))
        conn.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
#  KIT PARTS CRUD
# ══════════════════════════════════════════════════════════════════════════════

class CreateKitPart(BaseModel):
    name: str
    kit_name: Optional[str] = ""
    part_type: Optional[str] = ""
    description: Optional[str] = ""
    project: Optional[str] = ""
    subcategory: Optional[str] = ""
    resistance: Optional[str] = ""
    box_location: Optional[str] = ""
    glycerol_location: Optional[str] = ""
    source_url: Optional[str] = ""
    notes: Optional[str] = ""

class UpdateKitPart(BaseModel):
    name: Optional[str] = None
    kit_name: Optional[str] = None
    part_type: Optional[str] = None
    description: Optional[str] = None
    project: Optional[str] = None
    subcategory: Optional[str] = None
    resistance: Optional[str] = None
    box_location: Optional[str] = None
    glycerol_location: Optional[str] = None
    source_url: Optional[str] = None
    notes: Optional[str] = None
    stock_dna: Optional[str] = None
    stock_glycerol: Optional[str] = None
    stock_verified: Optional[str] = None

@router.get("/kit-parts")
def list_kit_parts():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM kit_parts ORDER BY name ASC").fetchall()
    return {"items": [dict(r) for r in rows]}

@router.post("/kit-parts")
def create_kit_part(body: CreateKitPart):
    now = datetime.utcnow().isoformat()
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO kit_parts (name, kit_name, part_type, description, project, subcategory, resistance, "
            "box_location, glycerol_location, source_url, notes, created) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (body.name, body.kit_name, body.part_type, body.description, body.project, body.subcategory,
             body.resistance, body.box_location, body.glycerol_location, body.source_url, body.notes, now))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM kit_parts WHERE id=?", (cur.lastrowid,)).fetchone())
    return row

@router.put("/kit-parts/{item_id}")
def update_kit_part(item_id: int, body: UpdateKitPart):
    with get_db() as conn:
        existing = conn.execute("SELECT * FROM kit_parts WHERE id=?", (item_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Kit part not found")
        fields = {k: v for k, v in body.dict().items() if v is not None}
        if not fields:
            return dict(existing)
        sets = ", ".join(f"{k}=?" for k in fields)
        conn.execute(f"UPDATE kit_parts SET {sets} WHERE id=?", (*fields.values(), item_id))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM kit_parts WHERE id=?", (item_id,)).fetchone())
    return row

@router.delete("/kit-parts/{item_id}")
def delete_kit_part(item_id: int):
    path = os.path.join(GB_DIR, f"kitpart_{item_id}.gb")
    if os.path.exists(path):
        os.remove(path)
    with get_db() as conn:
        conn.execute("DELETE FROM kit_parts WHERE id=?", (item_id,))
        conn.commit()
    return {"ok": True}

# ── Kit Part .gb file ────────────────────────────────────────────────────────

@router.post("/kit-parts/{item_id}/gb")
async def upload_kit_part_gb(item_id: int, file: UploadFile = File(...)):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM kit_parts WHERE id=?", (item_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Kit part not found")
    contents = await file.read()
    stored = os.path.join(GB_DIR, f"kitpart_{item_id}.gb")
    with open(stored, "wb") as f:
        f.write(contents)
    with get_db() as conn:
        conn.execute("UPDATE kit_parts SET gb_file=? WHERE id=?", (file.filename, item_id))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM kit_parts WHERE id=?", (item_id,)).fetchone())
    return row

@router.get("/kit-parts/{item_id}/gb")
def download_kit_part_gb(item_id: int):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM kit_parts WHERE id=?", (item_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Kit part not found")
    stored = os.path.join(GB_DIR, f"kitpart_{item_id}.gb")
    if not os.path.exists(stored):
        raise HTTPException(404, "No .gb file attached")
    return FileResponse(stored, filename=row["gb_file"] or f"kitpart_{item_id}.gb",
                        media_type="application/octet-stream")

@router.delete("/kit-parts/{item_id}/gb")
def delete_kit_part_gb(item_id: int):
    stored = os.path.join(GB_DIR, f"kitpart_{item_id}.gb")
    if os.path.exists(stored):
        os.remove(stored)
    with get_db() as conn:
        conn.execute("UPDATE kit_parts SET gb_file=NULL WHERE id=?", (item_id,))
        conn.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
#  PARTS CRUD
# ══════════════════════════════════════════════════════════════════════════════

class CreatePart(BaseModel):
    name: str
    description: Optional[str] = ""
    sequence: Optional[str] = ""
    project: Optional[str] = ""
    subcategory: Optional[str] = ""
    source_feature: Optional[str] = ""
    source_id: Optional[str] = ""
    part_type: Optional[str] = ""
    box_location: Optional[str] = ""
    glycerol_location: Optional[str] = ""
    notes: Optional[str] = ""

class UpdatePart(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    sequence: Optional[str] = None
    project: Optional[str] = None
    subcategory: Optional[str] = None
    source_feature: Optional[str] = None
    source_id: Optional[str] = None
    part_type: Optional[str] = None
    box_location: Optional[str] = None
    glycerol_location: Optional[str] = None
    notes: Optional[str] = None
    stock_dna: Optional[str] = None
    stock_glycerol: Optional[str] = None
    stock_verified: Optional[str] = None

def _calc_part_length(seq: str) -> int:
    if not seq:
        return 0
    return len(re.sub(r'[^ACGTacgt]', '', seq))

@router.get("/parts")
def list_parts():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM parts ORDER BY name ASC").fetchall()
    return {"items": [dict(r) for r in rows]}

@router.post("/parts")
def create_part(body: CreatePart):
    now = datetime.utcnow().isoformat()
    length = _calc_part_length(body.sequence or "")
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO parts (name, description, sequence, length, project, subcategory, source_feature, "
            "source_id, part_type, box_location, glycerol_location, notes, created) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (body.name, body.description, body.sequence, length, body.project, body.subcategory,
             body.source_feature, body.source_id, body.part_type,
             body.box_location, body.glycerol_location, body.notes, now))
        conn.commit()
        item_id = cur.lastrowid

        # Auto-generate .gb from sequence
        gb_file = ""
        if body.sequence and body.sequence.strip():
            gb_file = _generate_gb_for_part(item_id, body.name, body.sequence)
            if gb_file:
                conn.execute("UPDATE parts SET gb_file=? WHERE id=?", (gb_file, item_id))
                conn.commit()

        row = dict(conn.execute("SELECT * FROM parts WHERE id=?", (item_id,)).fetchone())
    return row

@router.put("/parts/{item_id}")
def update_part(item_id: int, body: UpdatePart):
    with get_db() as conn:
        existing = conn.execute("SELECT * FROM parts WHERE id=?", (item_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Part not found")
        fields = {k: v for k, v in body.dict().items() if v is not None}
        if not fields:
            return dict(existing)
        if "sequence" in fields:
            fields["length"] = _calc_part_length(fields["sequence"])
        sets = ", ".join(f"{k}=?" for k in fields)
        conn.execute(f"UPDATE parts SET {sets} WHERE id=?", (*fields.values(), item_id))
        conn.commit()

        # Regenerate .gb if sequence updated
        if "sequence" in fields and fields["sequence"].strip():
            name = fields.get("name") or dict(existing)["name"]
            gb_file = _generate_gb_for_part(item_id, name, fields["sequence"])
            if gb_file:
                conn.execute("UPDATE parts SET gb_file=? WHERE id=?", (gb_file, item_id))
                conn.commit()

        row = dict(conn.execute("SELECT * FROM parts WHERE id=?", (item_id,)).fetchone())
    return row

@router.delete("/parts/{item_id}")
def delete_part(item_id: int):
    path = os.path.join(GB_DIR, f"part_{item_id}.gb")
    if os.path.exists(path):
        os.remove(path)
    with get_db() as conn:
        conn.execute("DELETE FROM parts WHERE id=?", (item_id,))
        conn.commit()
    return {"ok": True}

# ── Part .gb file ────────────────────────────────────────────────────────────

@router.post("/parts/{item_id}/gb")
async def upload_part_gb(item_id: int, file: UploadFile = File(...)):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM parts WHERE id=?", (item_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Part not found")
    contents = await file.read()
    stored = os.path.join(GB_DIR, f"part_{item_id}.gb")
    with open(stored, "wb") as f:
        f.write(contents)
    extracted_seq = _extract_seq_from_gb(contents)
    length = _calc_part_length(extracted_seq) if extracted_seq else (dict(row).get("length") or 0)
    with get_db() as conn:
        if extracted_seq:
            conn.execute("UPDATE parts SET gb_file=?, sequence=?, length=? WHERE id=?",
                         (file.filename, extracted_seq, length, item_id))
        else:
            conn.execute("UPDATE parts SET gb_file=? WHERE id=?", (file.filename, item_id))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM parts WHERE id=?", (item_id,)).fetchone())
    return row

@router.get("/parts/{item_id}/gb")
def download_part_gb(item_id: int):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM parts WHERE id=?", (item_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Part not found")
    stored = os.path.join(GB_DIR, f"part_{item_id}.gb")
    if not os.path.exists(stored):
        raise HTTPException(404, "No .gb file attached")
    return FileResponse(stored, filename=row["gb_file"] or f"part_{item_id}.gb",
                        media_type="application/octet-stream")

@router.delete("/parts/{item_id}/gb")
def delete_part_gb(item_id: int):
    stored = os.path.join(GB_DIR, f"part_{item_id}.gb")
    if os.path.exists(stored):
        os.remove(stored)
    with get_db() as conn:
        conn.execute("UPDATE parts SET gb_file=NULL WHERE id=?", (item_id,))
        conn.commit()
    return {"ok": True}


def _generate_gb_for_part(item_id: int, name: str, sequence: str):
    """Generate a GenBank file for a part, same logic as gblocks."""
    if not sequence or not HAS_BIOPYTHON:
        return ""
    clean_seq = re.sub(r'[^ACGTacgtNn]', '', sequence)
    if not clean_seq:
        return ""
    try:
        from Bio.Seq import Seq
        from Bio.SeqRecord import SeqRecord
        from Bio.SeqFeature import SeqFeature, FeatureLocation
        from Bio import SeqIO as _SeqIO

        rec = SeqRecord(
            Seq(clean_seq.upper()),
            id=name[:10].replace(" ", "_"),
            name=name[:10].replace(" ", "_"),
            description=name,
            annotations={"molecule_type": "DNA", "topology": "linear"},
        )
        rec.features.append(SeqFeature(
            FeatureLocation(0, len(clean_seq)), type="source",
            qualifiers={"mol_type": ["other DNA"], "organism": ["synthetic construct"], "label": [name]},
        ))
        rec.features.append(SeqFeature(
            FeatureLocation(0, len(clean_seq)), type="misc_feature",
            qualifiers={"label": [name], "note": ["assembled part"]},
        ))
        gb_path = os.path.join(GB_DIR, f"part_{item_id}.gb")
        with open(gb_path, "w") as f:
            _SeqIO.write(rec, f, "genbank")
        return f"{name}.gb".replace(" ", "_")
    except Exception:
        return ""


# ══════════════════════════════════════════════════════════════════════════════
#  STORAGE BOXES CRUD
# ══════════════════════════════════════════════════════════════════════════════

class CreateBox(BaseModel):
    name: str
    rows: int = 9
    cols: int = 9
    box_type: str = "mixed"
    location: Optional[str] = ""

class UpdateBox(BaseModel):
    name: Optional[str] = None
    rows: Optional[int] = None
    cols: Optional[int] = None
    box_type: Optional[str] = None
    location: Optional[str] = None

class CellAssign(BaseModel):
    row: str
    col: int
    item_type: str
    item_id: int

class CellClear(BaseModel):
    row: str
    col: int

@router.get("/boxes")
def list_boxes():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM storage_boxes ORDER BY name ASC").fetchall()
    items = []
    for r in rows:
        d = dict(r)
        try:
            d["layout"] = json.loads(d["layout"] or "{}")
        except Exception:
            d["layout"] = {}
        items.append(d)
    return {"items": items}

@router.post("/boxes")
def create_box(body: CreateBox):
    now = datetime.utcnow().isoformat()
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO storage_boxes (name, rows, cols, box_type, location, layout, created, updated) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (body.name, body.rows, body.cols, body.box_type, body.location, "{}", now, now))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM storage_boxes WHERE id=?", (cur.lastrowid,)).fetchone())
    try:
        row["layout"] = json.loads(row["layout"] or "{}")
    except Exception:
        row["layout"] = {}
    return row

@router.put("/boxes/{box_id}")
def update_box(box_id: int, body: UpdateBox):
    now = datetime.utcnow().isoformat()
    with get_db() as conn:
        existing = conn.execute("SELECT * FROM storage_boxes WHERE id=?", (box_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Box not found")
        fields = {k: v for k, v in body.dict().items() if v is not None}
        if not fields:
            d = dict(existing)
            try:
                d["layout"] = json.loads(d["layout"] or "{}")
            except Exception:
                d["layout"] = {}
            return d
        fields["updated"] = now
        sets = ", ".join(f"{k}=?" for k in fields)
        conn.execute(f"UPDATE storage_boxes SET {sets} WHERE id=?", (*fields.values(), box_id))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM storage_boxes WHERE id=?", (box_id,)).fetchone())
    try:
        row["layout"] = json.loads(row["layout"] or "{}")
    except Exception:
        row["layout"] = {}
    return row

@router.delete("/boxes/{box_id}")
def delete_box(box_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM storage_boxes WHERE id=?", (box_id,))
        conn.commit()
    return {"ok": True}

@router.put("/boxes/{box_id}/cell")
def assign_cell(box_id: int, body: CellAssign):
    now = datetime.utcnow().isoformat()
    with get_db() as conn:
        existing = conn.execute("SELECT * FROM storage_boxes WHERE id=?", (box_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Box not found")
        try:
            layout = json.loads(existing["layout"] or "{}")
        except Exception:
            layout = {}
        key = f"{body.row}{body.col}"
        layout[key] = {"type": body.item_type, "id": body.item_id}
        conn.execute("UPDATE storage_boxes SET layout=?, updated=? WHERE id=?",
                     (json.dumps(layout), now, box_id))
        conn.commit()
    return {"ok": True, "layout": layout}

@router.delete("/boxes/{box_id}/cell")
def clear_cell(box_id: int, body: CellClear):
    now = datetime.utcnow().isoformat()
    with get_db() as conn:
        existing = conn.execute("SELECT * FROM storage_boxes WHERE id=?", (box_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Box not found")
        try:
            layout = json.loads(existing["layout"] or "{}")
        except Exception:
            layout = {}
        key = f"{body.row}{body.col}"
        layout.pop(key, None)
        conn.execute("UPDATE storage_boxes SET layout=?, updated=? WHERE id=?",
                     (json.dumps(layout), now, box_id))
        conn.commit()
    return {"ok": True, "layout": layout}


class MultiCellAssign(BaseModel):
    cells: list          # [{"row": "A", "col": 1}, {"row": "A", "col": 2}, ...]
    item_type: str
    item_id: int

@router.put("/boxes/{box_id}/multi-cell")
def assign_multi_cell(box_id: int, body: MultiCellAssign):
    now = datetime.utcnow().isoformat()
    with get_db() as conn:
        existing = conn.execute("SELECT * FROM storage_boxes WHERE id=?", (box_id,)).fetchone()
        if not existing:
            raise HTTPException(404, "Box not found")
        try:
            layout = json.loads(existing["layout"] or "{}")
        except Exception:
            layout = {}
        for c in body.cells:
            key = f"{c['row']}{c['col']}"
            layout[key] = {"type": body.item_type, "id": body.item_id}
        conn.execute("UPDATE storage_boxes SET layout=?, updated=? WHERE id=?",
                     (json.dumps(layout), now, box_id))
        conn.commit()
    return {"ok": True, "layout": layout, "count": len(body.cells)}


@router.get("/dna/box-positions/{item_type}/{item_id}")
def get_box_positions(item_type: str, item_id: int):
    """Reverse lookup: find all box cells containing a given item."""
    positions = []
    with get_db() as conn:
        rows = conn.execute("SELECT id, name, layout FROM storage_boxes").fetchall()
    for r in rows:
        try:
            layout = json.loads(r["layout"] or "{}")
        except Exception:
            continue
        for key, val in layout.items():
            if val.get("type") == item_type and val.get("id") == item_id:
                positions.append({"box_id": r["id"], "box_name": r["name"], "cell": key})
    return {"positions": positions}


# ══════════════════════════════════════════════════════════════════════════════
#  REINDEX (close gaps in naming)
# ══════════════════════════════════════════════════════════════════════════════

class ReindexRequest(BaseModel):
    table: str           # primers, plasmids, gblocks, kit_parts, parts
    prefix: str          # e.g. "MR", "pMR"
    start_num: int = 1
    execute: bool = False  # False = preview, True = apply

_REINDEX_TABLES = {
    "primers": "primers",
    "plasmids": "plasmids",
    "gblocks": "gblocks",
    "kit_parts": "kit_parts",
    "parts": "parts",
}

@router.post("/dna/reindex")
def reindex_names(body: ReindexRequest):
    table = _REINDEX_TABLES.get(body.table)
    if not table:
        raise HTTPException(400, f"table must be one of {list(_REINDEX_TABLES.keys())}")
    prefix = body.prefix.strip()
    if not prefix:
        raise HTTPException(400, "prefix is required")

    pattern = f"{prefix}%"
    with get_db() as conn:
        rows = conn.execute(
            f"SELECT id, name FROM {table} WHERE name LIKE ? ORDER BY name ASC",
            (pattern,)
        ).fetchall()

    if not rows:
        return {"changes": [], "message": "No items match that prefix."}

    # Parse numeric suffix, sort by number
    items = []
    for r in rows:
        name = r["name"]
        suffix = name[len(prefix):]
        # Extract leading digits from suffix
        digits = ""
        for ch in suffix:
            if ch.isdigit():
                digits += ch
            else:
                break
        if digits and suffix == digits:  # only pure prefix+number names
            items.append({"id": r["id"], "old_name": name, "old_num": int(digits)})

    items.sort(key=lambda x: x["old_num"])

    # Build rename plan
    changes = []
    next_num = body.start_num
    for item in items:
        new_name = f"{prefix}{next_num}"
        if new_name != item["old_name"]:
            changes.append({
                "id": item["id"],
                "old_name": item["old_name"],
                "new_name": new_name
            })
        next_num += 1

    if not body.execute:
        return {"changes": changes, "total_items": len(items), "preview": True}

    # Execute renames
    renamed = 0
    with get_db() as conn:
        for c in changes:
            conn.execute(f"UPDATE {table} SET name=? WHERE id=?", (c["new_name"], c["id"]))
            renamed += 1
        conn.commit()

    return {"changes": changes, "renamed": renamed, "preview": False}


# ══════════════════════════════════════════════════════════════════════════════
#  DNA LINK SETTINGS
# ══════════════════════════════════════════════════════════════════════════════

class DnaSettings(BaseModel):
    primer_prefix: str = ""
    plasmid_prefix: str = ""

@router.get("/dna/settings")
def get_dna_settings():
    with get_db() as conn:
        row = conn.execute("SELECT * FROM dna_settings WHERE id=1").fetchone()
    if not row:
        return {"primer_prefix": "", "plasmid_prefix": ""}
    return dict(row)

@router.post("/dna/settings")
def save_dna_settings(body: DnaSettings):
    now = datetime.utcnow().isoformat()
    with get_db() as conn:
        existing = conn.execute("SELECT * FROM dna_settings WHERE id=1").fetchone()
        if existing:
            conn.execute("UPDATE dna_settings SET primer_prefix=?, plasmid_prefix=? WHERE id=1",
                         (body.primer_prefix, body.plasmid_prefix))
        else:
            conn.execute("INSERT INTO dna_settings (id, primer_prefix, plasmid_prefix, created) VALUES (1,?,?,?)",
                         (body.primer_prefix, body.plasmid_prefix, now))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM dna_settings WHERE id=1").fetchone())
    return row


# ══════════════════════════════════════════════════════════════════════════════
#  IMPORT (CSV / XLSX) — extended for gblocks + kit_parts
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/import/upload")
async def upload_for_preview(file: UploadFile = File(...)):
    contents = await file.read()
    original_name = file.filename or "upload"
    ext = os.path.splitext(original_name)[1].lower()

    if ext not in (".csv", ".tsv", ".xlsx", ".xls"):
        raise HTTPException(400, "Unsupported file type. Upload .csv, .tsv, or .xlsx.")
    if ext in (".xlsx", ".xls") and not HAS_OPENPYXL:
        raise HTTPException(400, "Excel support requires openpyxl.")

    temp_id = uuid.uuid4().hex
    temp_path = os.path.join(UPLOAD_DIR, f"{temp_id}{ext}")
    with open(temp_path, "wb") as f:
        f.write(contents)

    try:
        headers, preview = _parse_preview(contents, ext)
    except Exception as e:
        os.remove(temp_path)
        raise HTTPException(400, f"Could not parse file: {e}")

    return {"temp_id": temp_id, "filename": original_name, "ext": ext,
            "headers": headers, "preview": preview}


class ImportRequest(BaseModel):
    temp_id: str
    ext: str
    filename: str = ""
    record_type: str                       # primer, plasmid, gblock, kit_part, part
    col_name: int
    col_sequence: Optional[int] = None
    col_use: Optional[int] = None
    col_box_number: Optional[int] = None
    col_box_location: Optional[int] = None
    col_glycerol_location: Optional[int] = None
    col_project: Optional[int] = None
    col_supplier: Optional[int] = None
    col_order_id: Optional[int] = None
    col_notes: Optional[int] = None
    col_kit_name: Optional[int] = None
    col_part_type: Optional[int] = None
    col_description: Optional[int] = None
    col_resistance: Optional[int] = None
    col_source_url: Optional[int] = None


@router.post("/import/execute")
def execute_import(body: ImportRequest):
    valid_types = ("primer", "plasmid", "gblock", "kit_part", "part")
    if body.record_type not in valid_types:
        raise HTTPException(400, f"record_type must be one of {valid_types}")

    temp_path = os.path.join(UPLOAD_DIR, f"{body.temp_id}{body.ext}")
    if not os.path.exists(temp_path):
        raise HTTPException(404, "Upload not found — please re-upload.")

    with open(temp_path, "rb") as f:
        contents = f.read()

    data_rows = _parse_data_rows(contents, body.ext)
    now = datetime.utcnow().isoformat()
    display_name = body.filename or os.path.basename(temp_path)

    inserted = 0
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO imports (filename, record_type, record_count, created) VALUES (?,?,?,?)",
            (display_name, body.record_type, 0, now))
        import_id = cur.lastrowid

        for row in data_rows:
            name = _cell(row, body.col_name)
            if not name:
                continue

            if body.record_type == "primer":
                _ensure_migrations()
                conn.execute(
                    "INSERT INTO primers (import_id, name, sequence, use, box_number, project, created) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (import_id, name, _cell(row, body.col_sequence),
                     _cell(row, body.col_use), _cell(row, body.col_box_number),
                     _cell(row, body.col_project), now))
            elif body.record_type == "plasmid":
                _ensure_migrations()
                conn.execute(
                    "INSERT INTO plasmids (import_id, name, use, box_location, glycerol_location, project, created) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (import_id, name, _cell(row, body.col_use),
                     _cell(row, body.col_box_location), _cell(row, body.col_glycerol_location),
                     _cell(row, body.col_project), now))
            elif body.record_type == "gblock":
                seq = _cell(row, body.col_sequence)
                length = _calc_gblock_length(seq)
                conn.execute(
                    "INSERT INTO gblocks (import_id, name, sequence, length, project, use, supplier, order_id, box_number, notes, created) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (import_id, name, seq, length,
                     _cell(row, body.col_project), _cell(row, body.col_use),
                     _cell(row, body.col_supplier) or "IDT", _cell(row, body.col_order_id),
                     _cell(row, body.col_box_number), _cell(row, body.col_notes), now))
            elif body.record_type == "kit_part":
                conn.execute(
                    "INSERT INTO kit_parts (import_id, name, kit_name, part_type, description, project, resistance, "
                    "box_location, glycerol_location, source_url, notes, created) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (import_id, name, _cell(row, body.col_kit_name), _cell(row, body.col_part_type),
                     _cell(row, body.col_description), _cell(row, body.col_project),
                     _cell(row, body.col_resistance), _cell(row, body.col_box_location),
                     _cell(row, body.col_glycerol_location), _cell(row, body.col_source_url),
                     _cell(row, body.col_notes), now))
            elif body.record_type == "part":
                seq = _cell(row, body.col_sequence)
                length = _calc_part_length(seq)
                conn.execute(
                    "INSERT INTO parts (import_id, name, description, sequence, length, project, subcategory, "
                    "part_type, box_location, glycerol_location, notes, created) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (import_id, name, _cell(row, body.col_description), seq, length,
                     _cell(row, body.col_project), "",
                     _cell(row, body.col_part_type), _cell(row, body.col_box_location),
                     _cell(row, body.col_glycerol_location), _cell(row, body.col_notes), now))
            inserted += 1

        conn.execute("UPDATE imports SET record_count=? WHERE id=?", (inserted, import_id))
        conn.commit()

    return {"import_id": import_id, "record_count": inserted, "record_type": body.record_type}


@router.get("/import/history")
def import_history():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM imports ORDER BY created DESC").fetchall()
    return {"items": [dict(r) for r in rows]}

@router.delete("/import/{import_id}")
def delete_import(import_id: int):
    with get_db() as conn:
        imp = conn.execute("SELECT * FROM imports WHERE id=?", (import_id,)).fetchone()
        if not imp:
            raise HTTPException(404, "Import not found")
        table_map = {"primer": "primers", "plasmid": "plasmids", "gblock": "gblocks", "kit_part": "kit_parts", "part": "parts"}
        table = table_map.get(imp["record_type"], "primers")
        conn.execute(f"DELETE FROM {table} WHERE import_id=?", (import_id,))
        conn.execute("DELETE FROM imports WHERE id=?", (import_id,))
        conn.commit()
    return {"ok": True}


# ── Parse helpers ────────────────────────────────────────────────────────────

def _cell(row, idx):
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return str(row[idx]).strip()

def _parse_preview(contents: bytes, ext: str):
    if ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            raise ValueError("Spreadsheet is empty")
        headers = [str(c) if c is not None else f"Column {i+1}" for i, c in enumerate(rows[0])]
        preview = [[str(c) if c is not None else "" for c in row] for row in rows[1:6]]
    else:
        text = contents.decode("utf-8-sig")
        delimiter = "\t" if ext == ".tsv" else ","
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        all_rows = list(reader)
        if not all_rows:
            raise ValueError("CSV is empty")
        headers = [c if c.strip() else f"Column {i+1}" for i, c in enumerate(all_rows[0])]
        preview = all_rows[1:6]
    return headers, preview

def _parse_data_rows(contents: bytes, ext: str):
    if ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[1:]
        wb.close()
        return [[str(c) if c is not None else "" for c in row] for row in rows]
    else:
        text = contents.decode("utf-8-sig")
        delimiter = "\t" if ext == ".tsv" else ","
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        return list(reader)[1:]
